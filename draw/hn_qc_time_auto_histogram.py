import matplotlib
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import seaborn as sns

# 读取 Excel 文件
exp1_file_path = r'C:\Users\penglab\Documents\金银铜\all_result_7.16.xlsx'  # 将 'example.xlsx' 替换为你的 Excel 文件路径
exp1_sheet_name = 'autoQC_V2'       # 将 'Sheet1' 替换为你的工作表名称
exp1_column = 'P'

exp2_file_path = r'C:\Users\penglab\Documents\金银铜\autoQC_数据生产流程.xlsx'
exp2_sheet_name = "流程一"
exp2_column = 'N'

def get_detect_time(file_path, sheet_name, column):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    column_data = []
    for cell in sheet[column][1:]:
        if cell.value is not None:
            column_data.append(float(cell.value))
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(column_data)
    return column_data


auto_time = get_detect_time(exp1_file_path, exp1_sheet_name, exp1_column)
auto_time = [time for time in auto_time if time < 200]
manual_time = [25 * 60 + 13, 15 * 60 + 25, 7 * 60 + 20, 8 * 60 + 22, 15 * 60, 31 * 60, 23 * 60, 12 * 60, 5 * 60 + 46,
               7 * 60 + 26, 25 * 60 + 30, 27 * 60 + 2]

data = [auto_time, manual_time]

# auto_relative_freq = [data.count(x) / len(data) for x in data]
matplotlib.use('TkAgg')
plt.figure(figsize=(10, 8))
plt.hist(auto_time, bins=20, edgecolor='black', alpha=0.7, density=True, label='histogram')

# 设置自定义的 x 轴刻度
x_ticks = np.arange(0, 61, 5)  # 定义刻度位置
# x_labels =   # 定义刻度标签
plt.xticks(ticks=x_ticks, fontsize=15)
plt.yticks(fontsize=15)

# 添加标题和标签
# plt.title('Frequency distribution of automatic QC inspect duration', fontsize=20, pad=30)
plt.xlabel('Proofreading time(s)', fontsize=19)
plt.ylabel('Density', fontsize=18)

# 调整布局以确保标签显示不被截断
plt.tight_layout()

# 移除图形边框
for spine in plt.gca().spines.values():
    if spine.spine_type in ['top', 'right']:
        spine.set_visible(False)

sns.kdeplot(auto_time, label='density')

# legend = plt.legend(loc='upper right', prop={'size': 14})  # 另一种方法，指定字体大小

# 设置 x 轴数据范围
plt.xlim(0, 60)
# 显示图形
plt.show()