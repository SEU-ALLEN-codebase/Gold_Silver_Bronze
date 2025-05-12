import matplotlib
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取 Excel 文件
exp1_file_path = r'C:\Users\10422\Documents\金银铜\all_result_7.16.xlsx'  # 将 'example.xlsx' 替换为你的 Excel 文件路径
exp1_sheet_name = 'autoQC_V2'       # 将 'Sheet1' 替换为你的工作表名称
exp1_column = 'P'

# exp2_file_path = r'C:\Users\penglab\Documents\金银铜\autoQC_数据生产流程.xlsx'
# exp2_sheet_name = "流程一"
# exp2_column = 'N'

def get_detect_time(file_path, sheet_name, column):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    column_data = []
    for cell in sheet[column][1:]:
        if cell.value is not None:
            column_data.append(float(cell.value))
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(column_data)
    return column_data

auto_time = get_detect_time(exp1_file_path, exp1_sheet_name, exp1_column)
auto_time = [time for time in auto_time if time < 200 and time > 5]
# manual_time = [25 * 60 + 13, 15 * 60 + 25, 7 * 60 + 20, 8 * 60 + 22, 15 * 60, 31 * 60, 23 * 60, 12 * 60, 5 * 60 + 46,
#                7 * 60 + 26, 25 * 60 + 30, 27 * 60 + 2]
manual_time = [1513, 925, 440, 502, 900, 1860, 1380, 720, 346, 446, 1530, 1622, 1165, 1162,
               915, 701, 507, 572, 429, 950, 967, 1135, 644, 449, 1307, 592, 1049, 1360, 1170,
               534, 1203, 1121, 726, 1060, 1277, 881, 716, 1014, 1153, 951, 597, 1239, 950,
               1310, 814, 1263, 577, 1343, 598, 628, 1049, 552, 435, 433, 1085, 1065, 1349,
               1349, 1352, 1115, 484, 973, 420, 792, 768, 791, 973, 462, 844, 1347, 536, 1021,
               736, 1325, 1205, 915, 868, 1303, 692, 1264, 1308, 451, 483, 771, 1374, 654, 1058,
               483, 949, 464, 498, 652, 994, 1034, 1148, 608, 1228, 811, 1295, 900]

auto_time = np.array(auto_time).astype(float)
manual_time = np.array(manual_time).astype(float)
data = [auto_time, manual_time]
auto_time += 1
manual_time += 1
log_auto_time = np.log10(auto_time)
log_maunal_time = np.log10(manual_time)
log_data = [log_auto_time, log_maunal_time]

# # 创建子图
# fig, axs = plt.subplots(1, 2, figsize=(12, 8))

# # 子图1
# axs[0].boxplot(auto_time)
# axs[0].set_xticks(ticks=[1], labels=['auto quality control'], fontsize=15)
# axs[0].set_title('Distribution of automatic QC inspect duration', fontsize=18, pad=35)
# axs[0].set_ylabel('Inspection time(s)', fontsize=15)

matplotlib.use('TkAgg')
plt.figure(figsize=(10, 8))

# 创建箱型图
plt.boxplot(log_data)

# 设置横坐标的标签
labels = ['Automatic QC', 'Manual']
plt.xticks(ticks=[1, 2], labels=labels, fontsize=15)

#添加标题和标签
plt.title('Time-consuming comparison of automatic QC inspection and manual inspection')
# plt.xlabel('Dataset')
# plt.yscale('log', base=10)  # 设置纵坐标为对数
plt.ylabel('Logarithm of inspect duration(s)', fontsize=15)

# 设置纵坐标标签字体大小
plt.yticks(fontsize=15)

# 调整布局以确保标签显示不被截断
plt.tight_layout()

# 移除图形边框
for spine in plt.gca().spines.values():
    if spine.spine_type in ['top', 'right']:
        spine.set_visible(False)

# 显示图形
plt.show()
