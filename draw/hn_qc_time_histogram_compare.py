import matplotlib
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import FuncFormatter, LogLocator, LogFormatter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import random
import seaborn as sns

# 读取 Excel 文件
from scipy import stats

exp1_file_path = r'C:\Users\penglab\Documents\金银铜\all_result_7.16.xlsx'  # 将 'example.xlsx' 替换为你的 Excel 文件路径
exp1_sheet_name = 'autoQC_V2'       # 将 'Sheet1' 替换为你的工作表名称
exp1_column = 'P'

# exp2_file_path = r'C:\Users\penglab\Documents\金银铜\autoQC_数据生产流程.xlsx'
# exp2_sheet_name = "流程一"
# exp2_column = 'N'

def get_detect_time(file_path, sheet_name, column):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    column_data = []
    for cell in sheet[column][1:]:
        if cell.value is not None:
            column_data.append(float(cell.value))
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(column_data)
    return column_data

auto_time = get_detect_time(exp1_file_path, exp1_sheet_name, exp1_column)
auto_time = [time for time in auto_time if time < 200 and time > 5]
# manual_time = [25 * 60 + 13, 15 * 60 + 25, 7 * 60 + 20, 8 * 60 + 22, 15 * 60, 31 * 60, 23 * 60, 12 * 60, 5 * 60 + 46,
#                7 * 60 + 26, 25 * 60 + 30, 27 * 60 + 2]
manual_time = [1513, 925, 440, 502, 900, 1860, 1380, 720, 346, 446, 1530, 1622, 1165, 1162,
               915, 701, 507, 572, 429, 950, 967, 1135, 644, 449, 1307, 592, 1049, 1360, 1170,
               534, 1203, 1121, 726, 1060, 1277, 881, 716, 1014, 1153, 951, 597, 1239, 950,
               1310, 814, 1263, 577, 1343, 598, 628, 1049, 552, 435, 433, 1085, 1065, 1349,
               1349, 1352, 1115, 484, 973, 420, 792, 768, 791, 973, 462, 844, 1347, 536, 1021,
               736, 1325, 1205, 915, 868, 1303, 692, 1264, 1308, 451, 483, 771, 1374, 654, 1058,
               483, 949, 464, 498, 652, 994, 1034, 1148, 608, 1228, 811, 1295, 900]

auto_time = np.array(auto_time).astype(float)
manual_time = np.array(manual_time).astype(float)
data = [np.copy(auto_time), np.copy(manual_time)]
# 计算标准差
std_devs = [np.std(group) for group in data]
# # 计算标准误差
# std_errs = [stats.sem(group) for group in data]
data_avg = [np.mean(auto_time), np.mean(manual_time)]
print(data_avg)

# 计算标准误差
std_errors = [np.std(group, ddof=1) / np.sqrt(len(group)) for group in data]

# 计算 95% 置信区间
confidence = 0.95
degrees_freedom_auto = len(auto_time) - 1
t_critical_auto = stats.t.ppf(confidence + (1-confidence)/2, degrees_freedom_auto)
degrees_freedom_manual = len(manual_time) - 1
t_critical_manual = stats.t.ppf(confidence + (1-confidence)/2, degrees_freedom_manual)

# 计算置信区间的误差条
errors = [t_critical_auto * std_errors[0], t_critical_manual * std_errors[1]]
print(errors)

auto_time += 1
manual_time += 1
log_auto_time = np.log2(auto_time)
log_manual_time = np.log2(manual_time)
log_data = [log_auto_time, log_manual_time]
log_data_avg = [np.mean(log_auto_time), np.mean(log_manual_time)]
# 计算标准差
log_std_devs = [np.std(group) for group in log_data]
# 计算标准误差
log_std_errs = [stats.sem(group) for group in log_data]

# auto_relative_freq = [data.count(x) / len(data) for x in data]

matplotlib.use('TkAgg')
plt.figure(figsize=(3, 6))

# distance_from_zero = 1.0  # 第一个柱子离零点的距离
# x_adjusted = x + distance_from_zero
types = ["            ", "      "]

# 设置初始位置和间距
bar_width = 0.12  # 柱子的宽度保持不变
spacing = 0.2    # 自定义间距

# 手动计算每个柱子的位置
x = np.arange(len(types)) * (bar_width + spacing)
print(x)

bars = plt.bar(x, data_avg, width=bar_width, yerr=errors, capsize=10, color=(250/255, 128/255, 144/255))

# # 设置自定义的 x 轴刻度
# x_ticks = np.arange(0, 61, 5)  # 定义刻度位置
# # x_labels =   # 定义刻度标签
plt.xticks(x, types, fontsize=20)
plt.yticks(fontsize=20)

# 设置 y 轴为对数刻度
plt.yscale('log', base=2)

# 自定义 y 轴刻度的标签格式，将其格式化为 2 的次方形式
def format_func(value, tick_number):
    if value == 0:
        return "0"
    exponent = int(np.log2(value))  # 获取以2为底的指数
    # if exponent <= 10:
    #     return f'$2^{{{exponent}}}$'
    #     # 当指数较大时，使用常规格式
    # else:
    return f'$2^{{{exponent}}}$'

# 设置 y 轴的主要刻度为 2 的次方形式，并且标签显示为 2 的次方
plt.gca().yaxis.set_major_locator(LogLocator(base=2))  # 设置主要刻度为2的次方
plt.gca().yaxis.set_major_formatter(FuncFormatter(format_func))  # 标签显示为2的次方

#添加标题和标签
# plt.title('Time-consuming comparison of automatic QC inspection and manual inspection', fontsize=21, pad=25)
# plt.xlabel('Dataset')
# plt.yscale('log', base=10)  # 设置纵坐标为对数
plt.ylabel('Check time(s)', fontsize=20)

plt.xlim([-0.2, 0.44])
plt.ylim(2**4, 2**10 + 100)
# 调整布局以确保标签显示不被截断
plt.tight_layout()

# 移除图形边框
for spine in plt.gca().spines.values():
    if spine.spine_type in ['top', 'right']:
        spine.set_visible(False)

# 保存为 SVG 格式
plt.savefig(r'C:\Users\penglab\Documents\金银铜\figure_svg\hn_qc_time_compare.svg', format='svg')

# 显示图形
plt.show()