import matplotlib
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import seaborn as sns
from openpyxl.styles import PatternFill

# 读取 Excel 文件
exp1_file_path = r'C:\Users\penglab\Documents\金银铜\all_result_7.16.xlsx'  # 将 'example.xlsx' 替换为你的 Excel 文件路径
exp1_sheet_name = 'autoQC_V2'       # 将 'Sheet1' 替换为你的工作表名称
exp1_column = 'P'

# exp2_file_path = r'C:\Users\penglab\Documents\金银铜\autoQC_数据生产流程.xlsx'
# exp2_sheet_name = "流程一"
# exp2_column = 'N'

def get_detect_time(file_path, sheet_name, column):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    column_data = []
    for cell in sheet[column][1:]:
        if cell.value is not None:
            column_data.append(float(cell.value))
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(column_data)
    return column_data

# auto_time = get_detect_time(exp1_file_path, exp1_sheet_name, exp1_column)
# auto_time = [time for time in auto_time if time < 200 and time > 5]
# manual_time = [25 * 60 + 13, 15 * 60 + 25, 7 * 60 + 20, 8 * 60 + 22, 15 * 60, 31 * 60, 23 * 60, 12 * 60, 5 * 60 + 46,
#                7 * 60 + 26, 25 * 60 + 30, 27 * 60 + 2]
manual_time = [1513, 925, 440, 502, 900, 1860, 1380, 720, 346, 446, 1530, 1622, 1165, 1162,
               915, 701, 507, 572, 429, 950, 967, 1135, 644, 449, 1307, 592, 1049, 1360, 1170,
               534, 1203, 1121, 726, 1060, 1277, 881, 716, 1014, 1153, 951, 597, 1239, 950,
               1310, 814, 1263, 577, 1343, 598, 628, 1049, 552, 435, 433, 1085, 1065, 1349,
               1349, 1352, 1115, 484, 973, 420, 792, 768, 791, 973, 462, 844, 1347, 536, 1021,
               736, 1325, 1205, 915, 868, 1303, 692, 1264, 1308, 451, 483, 771, 1374, 654, 1058,
               483, 949, 464, 498, 652, 994, 1034, 1148, 608, 1228, 811, 1295, 900]

manual_time= np.array(manual_time).astype(float)
manual_time_copy = manual_time.copy()
manual_time_copy += 1
log_manual_time = np.log2(manual_time_copy)

data = {
    'Category': ['Manual'] * len(manual_time),
    'Value': manual_time
}
df = pd.DataFrame(data)

# # 创建子图
# fig, axs = plt.subplots(1, 2, figsize=(12, 8))

# # 子图1
# axs[0].boxplot(auto_time)
# axs[0].set_xticks(ticks=[1], labels=['auto quality control'], fontsize=15)
# axs[0].set_title('Distribution of automatic QC inspect duration', fontsize=18, pad=35)
# axs[0].set_ylabel('Inspection time(s)', fontsize=15)

matplotlib.use('TkAgg')
plt.figure(figsize=(6, 8))

# # 创建箱型图
# plt.boxplot(log_auto_time)

# 绘制小提琴图
sns.violinplot(x='Category', y='Value', data=df, bw=0.1, scale='count', palette='Set2')

# 计算每个类别的四分位数
quartiles = df.groupby('Category')['Value'].quantile([0.25, 0.5, 0.75]).unstack()
quartiles.columns = ['Q1', 'Median', 'Q3']

# # 绘制四分位线
# for i, (category, row) in enumerate(quartiles.iterrows()):
#     plt.plot([i - 0.4, i + 0.4], [row['Q1'], row['Q1']], color='blue', linestyle='--', label='Q1' if i == 0 else "")
#     plt.plot([i - 0.4, i + 0.4], [row['Median'], row['Median']], color='green', linestyle='-', label='Median' if i == 0 else "")
#     plt.plot([i - 0.4, i + 0.4], [row['Q3'], row['Q3']], color='red', linestyle='--', label='Q3' if i == 0 else "")

# 添加图例
# plt.legend()
# # 添加均值线
# mean_values = df.groupby('Category')['Value'].mean().values
# for i, mean in enumerate(mean_values):
#     plt.axhline(mean, color='red', linestyle='--', xmin=i/2, xmax=(i+1)/2)
# # 叠加散点图
# sns.swarmplot(x='Category', y='Value', data=df, color='k', alpha=0.6)
# sns.stripplot(x='Category', y='Value', data=df, jitter=True)

# # 设置横坐标的标签
# labels = ['Automatic QC']
# plt.xticks(ticks=[1], labels=labels, fontsize=15)

# #添加标题和标签
# plt.title('Time-consuming comparison of automatic QC inspection and manual inspection')
plt.xlabel('')
plt.xticks(fontsize=21)
# plt.yscale('log', base=10)  # 设置纵坐标为对数
plt.ylabel('check time(s)', fontsize=21)

# 设置纵坐标标签字体大小
plt.yticks(fontsize=18)

# 调整布局以确保标签显示不被截断
plt.tight_layout()

# 移除图形边框
for spine in plt.gca().spines.values():
    if spine.spine_type in ['top', 'right']:
        spine.set_visible(False)

# 显示图形
plt.show()
