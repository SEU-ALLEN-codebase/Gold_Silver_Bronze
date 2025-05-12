import pandas as pd
import os, shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取 Excel 文件
file_path = r'C:\Users\penglab\Documents\金银铜\all_result_7.16.xlsx'  # 将 'example.xlsx' 替换为你的 Excel 文件路径
sheet_name = 'autoQC_V1'       # 将 'Sheet1' 替换为你的工作表名称
columns = ['C', 'E', 'F', 'I', 'J']
missing_column = ['G']
error_type_dict = {
    'C': 'Multifurcation',
    'E': 'Loop',
    'F': 'Angle Error',
    'J': 'Floating Branch',
    'I': 'Overlapping Branch',
    'G': 'Missing'
}

def count_cells_with_parentheses_or_nonwhite_bg(file_path, sheet_name, columns):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    # 初始化计数器
    counts = {col: 0 for col in columns}

    for col in columns:
        for cell in sheet[col][1:]:
            # 确保单元格值不是 None 或空值
            cell_value = str(cell.value) if cell.value is not None else ''

            # 检查单元格内容是否带括号
            if '(' in str(cell.value) or '（' in str(cell.value):
                print(cell_value)
                index = cell_value.find('（')
                if index == -1:
                    index = cell_value.index('(')
                if cell_value[index - 1] == '对':
                    counts[col] += int(cell_value[index - 2])
                else:
                    counts[col] += int(cell_value[index - 1])
                continue

            # 检查单元格背景颜色是否为白色
            fill = cell.fill
            if fill and isinstance(fill, PatternFill):
                fgColor = fill.fgColor.rgb if fill.fgColor else None
                if fgColor and fgColor != '00000000' and fgColor != 'FFFFFFFF':  # Assuming '00000000' and 'FFFFFFFF' as white
                    counts[col] += int(str(cell.value))

    return counts

print("各种错误 TP + FP 的数量:")
# 统计
result = count_cells_with_parentheses_or_nonwhite_bg(file_path, sheet_name, columns)
print(result)
