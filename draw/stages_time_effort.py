import matplotlib
from openpyxl import load_workbook
from scipy import stats
import matplotlib.pyplot as plt
import numpy as np

excel_path = r'C:\Users\penglab\Documents\金银铜\autoQC_数据生产流程_20241018.xlsx'

bronze_sheet_name = r'铜标准流程三'
silver_sheet_name = r'银标准检查与修改'
gold_sheet_name = r'金标准检查与修改'
bronze_manual_check_time_column = 'O'
bronze_manual_modify_time_column = 'P'
human_bronze_row_start = 50
human_bronze_row_end = 92
mouse_bronze_row_start = 94
mouse_bronze_row_end = 123
human_bronze_row_exclude = 61

silver_manual_time = 'H'
human_silver_check_row_start = 3
human_silver_check_row_end = 44
mouse_silver_check_row_start = 46
mouse_silver_check_row_end = 75
human_silver_modify_row_start = 82
human_silver_modify_row_end = 123
mouse_silver_modify_row_start = 125
mouse_silver_modify_row_end = 155
mouse_silver_row_exclude = 130

human_gold_part1_check_time_column = 'I'
human_gold_part1_modify_time_column = 'I'
gold_part2_check_time_column = 'K'
gold_part2_modify_time_column = 'O'
human_gold_part1_check_row_start = 3
human_gold_part1_check_row_end = 32
human_gold_part1_modify_row_start = 37
human_gold_part1_modify_row_end = 66
human_gold_part2_check_modify_row_start = 70
human_gold_part2_check_modify_row_end = 81
mouse_gold_check_modify_row_start = 83
mouse_gold_check_modify_row_end = 112

removed_human_neuron_number = ['03764', '05578', '06019']
name_column = 'A'

human_bronze_time, human_silver_time, human_gold_time = [], [], []
mouse_bronze_time, mouse_silver_time, mouse_gold_time = [], [], []

def get_time(file_path, sheet_name, column, row, time_list, is_human=True, removed_row=None):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    row_start = row[0]
    row_end = row[1]
    for i, cell in enumerate(sheet[column][row_start-1:row_end]):
        if row_start + i == removed_row:
            continue
        if cell.value is not None and cell.value != '':
            if is_human:
                image_number = sheet[name_column][row_start-1 + i].value.split('_')[0]
                if image_number in removed_human_neuron_number:
                    continue
            time_str = str(cell.value).strip()
            time = 0
            if 'min' in time_str and 's' in time_str:
                left, right = time_str.split('min')
                # print(left, right)
                min_val = float(left)
                time += min_val * 60
                # print(right.split('s')[0])
                sec_val = float(right.split('s')[0])
                time += sec_val
            elif 'min' in time_str:
                left = time_str.split('min')[0]
                min_val = float(left)
                time += min_val * 60
            time_list.append(time)

    print(time_list)

human_bronze_check_time, human_bronze_modify_time = [], []
get_time(excel_path, bronze_sheet_name, bronze_manual_check_time_column, [human_bronze_row_start, human_bronze_row_end], human_bronze_check_time, True, 61)
get_time(excel_path, bronze_sheet_name, bronze_manual_modify_time_column, [human_bronze_row_start, human_bronze_row_end], human_bronze_modify_time, True, 61)
for i in range(len(human_bronze_check_time)):
    human_bronze_time.append(human_bronze_check_time[i] + human_bronze_modify_time[i])

mouse_bronze_check_time, mouse_bronze_modify_time = [], []
get_time(excel_path, bronze_sheet_name, bronze_manual_check_time_column, [mouse_bronze_row_start, mouse_bronze_row_end], mouse_bronze_check_time, False)
get_time(excel_path, bronze_sheet_name, bronze_manual_modify_time_column, [mouse_bronze_row_start, mouse_bronze_row_end], mouse_bronze_modify_time, False)
for i in range(len(mouse_bronze_check_time)):
    mouse_bronze_time.append(mouse_bronze_check_time[i] + mouse_bronze_modify_time[i])

human_silver_check_time, human_silver_modify_time = [], []
get_time(excel_path, silver_sheet_name, silver_manual_time, [human_silver_check_row_start, human_silver_check_row_end], human_silver_check_time, True)
get_time(excel_path, silver_sheet_name, silver_manual_time, [human_silver_modify_row_start, human_silver_modify_row_end], human_silver_modify_time, True)
for i in range(len(human_silver_check_time)):
    human_silver_time.append(human_silver_check_time[i] + human_silver_modify_time[i])

mouse_silver_check_time, mouse_silver_modify_time = [], []
get_time(excel_path, silver_sheet_name, silver_manual_time, [mouse_silver_check_row_start, mouse_silver_check_row_end], mouse_silver_check_time, False)
get_time(excel_path, silver_sheet_name, silver_manual_time, [mouse_silver_modify_row_start, mouse_silver_modify_row_end], mouse_silver_modify_time, False, 130)
for i in range(len(mouse_silver_check_time)):
    mouse_silver_time.append(mouse_silver_check_time[i] + mouse_silver_modify_time[i])

human_gold_check_time_part1, human_gold_modify_time_part1 = [], []
human_gold_check_time_part2, human_gold_modify_time_part2 = [], []
mouse_gold_check_time, mouse_gold_modify_time = [], []
get_time(excel_path, gold_sheet_name, human_gold_part1_check_time_column, [human_gold_part1_check_row_start, human_gold_part1_check_row_end], human_gold_check_time_part1, True)
get_time(excel_path, gold_sheet_name, human_gold_part1_modify_time_column, [human_gold_part1_modify_row_start, human_gold_part1_modify_row_end], human_gold_modify_time_part1, True)
get_time(excel_path, gold_sheet_name, gold_part2_check_time_column, [human_gold_part2_check_modify_row_start, human_gold_part2_check_modify_row_end], human_gold_check_time_part2, True)
get_time(excel_path, gold_sheet_name, gold_part2_modify_time_column, [human_gold_part2_check_modify_row_start, human_gold_part2_check_modify_row_end], human_gold_modify_time_part2, True)
get_time(excel_path, gold_sheet_name, gold_part2_check_time_column, [mouse_gold_check_modify_row_start, mouse_gold_check_modify_row_end], mouse_gold_check_time, False)
get_time(excel_path, gold_sheet_name, gold_part2_modify_time_column, [mouse_gold_check_modify_row_start, mouse_gold_check_modify_row_end], mouse_gold_modify_time, False)
for i in range(len(human_gold_check_time_part1)):
    human_gold_time.append(human_gold_check_time_part1[i] + human_gold_modify_time_part1[i])
for i in range(len(human_gold_check_time_part2)):
    human_gold_time.append(human_gold_check_time_part2[i] + human_gold_modify_time_part2[i])
for i in range(len(mouse_gold_check_time)):
    mouse_gold_time.append(mouse_gold_check_time[i] + mouse_gold_modify_time[i])

print(len(human_bronze_time), len(human_silver_time), len(human_gold_time))
assert len(human_bronze_time) == len(human_silver_time) == len(human_gold_time)
assert len(mouse_bronze_time) == len(mouse_silver_time) == len(mouse_gold_time)

human_bronze_time = np.array(human_bronze_time).astype(float)
human_silver_time = np.array(human_silver_time).astype(float)
human_gold_time = np.array(human_gold_time).astype(float)
human_time_data = [np.copy(human_bronze_time), np.copy(human_silver_time), np.copy(human_gold_time)]

mouse_bronze_time = np.array(mouse_bronze_time).astype(float)
mouse_silver_time = np.array(mouse_silver_time).astype(float)
mouse_gold_time = np.array(mouse_gold_time).astype(float)
mouse_time_data = [np.copy(mouse_bronze_time), np.copy(mouse_silver_time), np.copy(mouse_gold_time)]

# 计算标准差
std_human_time_devs = [np.std(group) for group in human_time_data]
# 计算标准误差
# std_errs = [stats.sem(group) for group in data]
data_human_time_avg = [46, np.mean(human_bronze_time), np.mean(human_silver_time), np.mean(human_gold_time)]
# data_avg = [np.mean(mouse_bronze_time), np.mean(mouse_silver_time), np.mean(mouse_gold_time)]
std_mouse_time_devs = [np.std(group) for group in mouse_time_data]
data_mouse_time_avg = [55, np.mean(mouse_bronze_time), np.mean(mouse_silver_time), np.mean(mouse_gold_time)]

log_data_human_time_avg = [np.log2(data) for data in data_human_time_avg]
log_data_mouse_time_avg = [np.log2(data) for data in data_mouse_time_avg]

data_human_time_avg_accumulate = []
data_mouse_time_avg_accumulate = []
time_sum = 0
for i in range(len(data_human_time_avg)):
    time_sum += data_human_time_avg[i]
    data_human_time_avg_accumulate.append(time_sum)
time_sum = 0
for i in range(len(data_mouse_time_avg)):
    time_sum += data_mouse_time_avg[i]
    data_mouse_time_avg_accumulate.append(time_sum)

data_human_time_avg_accumulate = np.array(data_human_time_avg_accumulate).astype(float)
data_mouse_time_avg_accumulate = np.array(data_mouse_time_avg_accumulate).astype(float)

log_data_human_time_avg_accumulate = [np.log2(time) for time in data_human_time_avg_accumulate]
log_data_mouse_time_avg_accumulate = [np.log2(time) for time in data_mouse_time_avg_accumulate]

sqrt_data_human_time_avg_accumulate = np.sqrt(data_human_time_avg_accumulate)
sqrt_data_mouse_time_avg_accumulate = np.sqrt(data_mouse_time_avg_accumulate)

excel_path2 = r'C:\Users\penglab\Documents\金银铜\autoQC_数据生产流程_for_synthetic.xlsx'
bronze_sheet_name = r'铜标准流程三'
silver_sheet_name = r'银标准检查与修改'
gold_sheet_name = r'金标准检查与修改'
synthetic_bronze_manual_check_time_column = 'O'
synthetic_bronze_manual_modify_time_column = 'P'
synthetic_bronze_row_start = 125
synthetic_bronze_row_end = 139

synthetic_silver_manual_time = 'F'
synthetic_silver_check_row_start = 161
synthetic_silver_check_row_end = 175
synthetic_silver_modify_row_start = 179
synthetic_silver_modify_row_end = 193

synthetic_gold_check_time_column = 'K'
synthetic_gold_modify_time_column = 'O'
synthetic_gold_manual_row_start = 117
synthetic_gold_manual_row_end = 131

synthetic_bronze_time = []
synthetic_silver_time = []
synthetic_gold_time = []

synthetic_bronze_check_time, synthetic_bronze_modify_time = [], []
get_time(excel_path2, bronze_sheet_name, synthetic_bronze_manual_check_time_column, [synthetic_bronze_row_start, synthetic_bronze_row_end], synthetic_bronze_check_time, False)
get_time(excel_path2, bronze_sheet_name, synthetic_bronze_manual_modify_time_column, [synthetic_bronze_row_start, synthetic_bronze_row_end], synthetic_bronze_modify_time, False)
for i in range(len(synthetic_bronze_check_time)):
    synthetic_bronze_time.append(synthetic_bronze_check_time[i] + synthetic_bronze_modify_time[i])

synthetic_silver_check_time, synthetic_silver_modify_time = [], []
get_time(excel_path2, silver_sheet_name, synthetic_silver_manual_time, [synthetic_silver_check_row_start, synthetic_silver_check_row_end], synthetic_silver_check_time, False)
get_time(excel_path2, silver_sheet_name, synthetic_silver_manual_time, [synthetic_silver_modify_row_start, synthetic_silver_modify_row_end], synthetic_silver_modify_time, False)
for i in range(len(synthetic_silver_check_time)):
    synthetic_silver_time.append(synthetic_silver_check_time[i] +synthetic_silver_modify_time[i])

synthetic_gold_check_time, synthetic_gold_modify_time = [], []
get_time(excel_path2, gold_sheet_name, synthetic_gold_check_time_column, [synthetic_gold_manual_row_start, synthetic_gold_manual_row_end], synthetic_gold_check_time, False)
get_time(excel_path2, gold_sheet_name, synthetic_gold_modify_time_column, [synthetic_gold_manual_row_start, synthetic_gold_manual_row_end], synthetic_gold_modify_time, False)
for i in range(len(synthetic_gold_check_time)):
    synthetic_gold_time.append(synthetic_gold_check_time[i] + synthetic_gold_modify_time[i])

assert len(synthetic_bronze_time) == len(synthetic_silver_time) == len(synthetic_gold_time)

synthetic_bronze_time = np.array(synthetic_bronze_time).astype(float)
synthetic_silver_time = np.array(synthetic_silver_time).astype(float)
synthetic_gold_time = np.array(synthetic_gold_time).astype(float)
synthetic_time_data = [np.copy(synthetic_bronze_time), np.copy(synthetic_silver_time), np.copy(synthetic_gold_time)]

data_synthetic_time_avg = [52, np.mean(synthetic_bronze_time), np.mean(synthetic_silver_time), np.mean(synthetic_gold_time)]
data_synthetic_time_avg_accumulate = []
time_sum = 0
for i in range(len(data_synthetic_time_avg)):
    time_sum += data_synthetic_time_avg[i]
    data_synthetic_time_avg_accumulate.append(time_sum)

data_synthetic_time_avg_accumulate = np.array(data_synthetic_time_avg_accumulate).astype(float)


# data_human_time_avg_accumulate_norm = (data_human_time_avg_accumulate - np.min(data_human_time_avg_accumulate)) / (np.max(data_human_time_avg_accumulate) - np.min(data_human_time_avg_accumulate))
# data_mouse_time_avg_accumulate_norm = (data_mouse_time_avg_accumulate - np.min(data_mouse_time_avg_accumulate)) / (np.max(data_mouse_time_avg_accumulate) - np.min(data_mouse_time_avg_accumulate))
# print('data_human_time_avg_accumulate_norm: ', data_human_time_avg_accumulate_norm)
# print('data_mouse_time_avg_accumulate_norm: ', data_mouse_time_avg_accumulate_norm)

exp1_file_path = r'D:\source\PythonProjects\Gold_Silver_Bronze\experiments\eval_human_all.xlsx'  # 将 'example.xlsx' 替换为你的 Excel 文件路径
exp1_sheet_name = 'Sheet1'       # 将 'Sheet1' 替换为你的工作表名称
mouse_18454_file_path = r'D:\source\PythonProjects\Gold_Silver_Bronze\experiments\mouse_18454_eval.xlsx'
mouse_7_file_path = r'D:\source\PythonProjects\Gold_Silver_Bronze\experiments\7_mouse_eval.xlsx'
mouse_18454_sheet_name = 'mouse_18454_eval'
mouse_7_sheet_name = '7_mouse_eval'
mouse_all_file_path = r'D:\source\PythonProjects\Gold_Silver_Bronze\experiments\eval_mouse_all.xlsx'
synthetic_15_file_path = r'D:\source\PythonProjects\Gold_Silver_Bronze\experiments\15_synthetic_eval.xlsx'
synthetic_15_sheet_name = '15_synthetic_eval'

name_column = 'A'
auto_column = 'B'
bronze_column = 'C'
silver_column = 'D'

auto_recall_column = 'E'
bronze_recall_column = 'F'
silver_recall_column = 'G'

synthetic_auto_precision_column = 'B'
synthetic_bronze_precision_column = 'C'
synthetic_silver_precision_column = 'D'
synthetic_gold_precision_column = 'E'

synthetic_auto_recall_column = 'F'
synthetic_bronze_recall_column = 'G'
synthetic_silver_recall_column = 'H'
synthetic_gold_recall_column = 'I'

removed_human_neuron_number = ['03764', '05578', '06019']

matplotlib.use('TkAgg')

def get_precision(file_path, sheet_name, auto_column, bronze_column, silver_column, is_human=True):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    auto_data = []
    count = 0
    for i, cell in enumerate(sheet[auto_column][1:]):
        if cell.value is not None and cell.value != '':
            if is_human:
                image_number = sheet[name_column][i + 1].value.split('_')[0]
                if image_number in removed_human_neuron_number:
                    continue
            count += 1
            auto_data.append(float(cell.value))
        else:
            break
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(auto_data)

    bronze_data = []
    for i, cell in enumerate(sheet[bronze_column][1:]):
        if cell.value is not None and cell.value != '':
            if is_human:
                image_number = sheet[name_column][i + 1].value.split('_')[0]
                if image_number in removed_human_neuron_number:
                    continue
            bronze_data.append(float(cell.value))
        else:
            break
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(bronze_data)

    silver_data = []
    for i, cell in enumerate(sheet[silver_column][1:]):
        if cell.value is not None and cell.value != '':
            if is_human:
                image_number = sheet[name_column][i + 1].value.split('_')[0]
                if image_number in removed_human_neuron_number:
                    continue
            silver_data.append(float(cell.value))
        else:
            break
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(silver_data)

    gold_data = [1.0] * count

    return auto_data, bronze_data, silver_data, gold_data

auto_data, bronze_data, silver_data, gold_data = get_precision(exp1_file_path, exp1_sheet_name, auto_column, bronze_column, silver_column, True)
auto_recall_data, bronze_recall_data, silver_recall_data, gold_recall_data = get_precision(exp1_file_path, exp1_sheet_name, auto_recall_column, bronze_recall_column, silver_recall_column, True)

assert len(auto_data) == len(bronze_data) == len(silver_data) == len(gold_data)

# auto_data = [data * 100 for data in auto_data]
# bronze_data = [data * 100 for data in bronze_data]
# silver_data = [data * 100 for data in silver_data]
# gold_data = [data * 100 for data in gold_data]

auto_data = np.array(auto_data).astype(float)
bronze_data = np.array(bronze_data).astype(float)
silver_data = np.array(silver_data).astype(float)
gold_data = np.array(gold_data).astype(float)
data = [np.copy(auto_data), np.copy(bronze_data), np.copy(silver_data), np.copy(gold_data)]

data_avg_precision = [np.mean(auto_data), np.mean(bronze_data), np.mean(silver_data), np.mean(gold_data)]
print(data_avg_precision)

human_incr_data_avg_precision = [np.mean(auto_data), np.mean(bronze_data) - np.mean(auto_data), np.mean(silver_data) - np.mean(bronze_data), np.mean(gold_data) - np.mean(silver_data)]

# auto_recall_data = [data * 100 for data in auto_recall_data]
# bronze_recall_data = [data * 100 for data in bronze_recall_data]
# silver_recall_data = [data * 100 for data in silver_recall_data]
# gold_recall_data = [data * 100 for data in gold_recall_data]

auto_recall_data = np.array(auto_recall_data).astype(float)
bronze_recall_data = np.array(bronze_recall_data).astype(float)
silver_recall_data = np.array(silver_recall_data).astype(float)
gold_recall_data = np.array(gold_recall_data).astype(float)
data_recall = [np.copy(auto_recall_data), np.copy(bronze_recall_data), np.copy(silver_recall_data), np.copy(gold_recall_data)]

data_avg_recall = [np.mean(auto_recall_data), np.mean(bronze_recall_data), np.mean(silver_recall_data), np.mean(gold_recall_data)]
print(data_avg_recall)

human_incr_data_avg_recall = [np.mean(auto_recall_data), np.mean(bronze_recall_data) - np.mean(auto_recall_data), np.mean(silver_recall_data) - np.mean(bronze_recall_data), np.mean(gold_recall_data) - np.mean(silver_recall_data)]

mouse_auto_data, mouse_bronze_data, mouse_silver_data, mouse_gold_data = get_precision(mouse_all_file_path, mouse_7_sheet_name, auto_column, bronze_column, silver_column, False)
mouse_auto_recall_data, mouse_bronze_recall_data, mouse_silver_recall_data, mouse_gold_recall_data = get_precision(mouse_all_file_path, mouse_7_sheet_name, auto_recall_column, bronze_recall_column, silver_recall_column, False)

stages = ['Auto', 'Bronze', 'Silver', 'Gold']

assert len(mouse_auto_data) == len(mouse_bronze_data) == len(mouse_silver_data) == len(mouse_gold_data)

# mouse_auto_data = [data * 100 for data in mouse_auto_data]
# mouse_bronze_data = [data * 100 for data in mouse_bronze_data]
# mouse_silver_data = [data * 100 for data in mouse_silver_data]
# mouse_gold_data = [data * 100 for data in mouse_gold_data]

mouse_auto_data = np.array(mouse_auto_data).astype(float)
mouse_bronze_data = np.array(mouse_bronze_data).astype(float)
mouse_silver_data = np.array(mouse_silver_data).astype(float)
mouse_gold_data = np.array(mouse_gold_data).astype(float)
mouse_data = [np.copy(mouse_auto_data), np.copy(mouse_bronze_data), np.copy(mouse_silver_data), np.copy(mouse_gold_data)]

# 计算标准误差
# std_errs = [stats.sem(group) for group in data]
mouse_data_avg_precision = [np.mean(mouse_auto_data), np.mean(mouse_bronze_data), np.mean(mouse_silver_data), np.mean(mouse_gold_data)]

mouse_incr_data_avg_precision = [np.mean(mouse_auto_data), np.mean(mouse_bronze_data) - np.mean(mouse_auto_data), np.mean(mouse_silver_data) - np.mean(mouse_bronze_data), np.mean(mouse_gold_data) - np.mean(mouse_silver_data)]

print(mouse_data_avg_precision)

# mouse_auto_recall_data = [data * 100 for data in mouse_auto_recall_data]
# mouse_bronze_recall_data = [data * 100 for data in mouse_bronze_recall_data]
# mouse_silver_recall_data = [data * 100 for data in mouse_silver_recall_data]
# mouse_gold_recall_data = [data * 100 for data in mouse_gold_recall_data]

mouse_auto_recall_data = np.array(mouse_auto_recall_data).astype(float)
mouse_bronze_recall_data = np.array(mouse_bronze_recall_data).astype(float)
mouse_silver_recall_data = np.array(mouse_silver_recall_data).astype(float)
mouse_gold_recall_data = np.array(mouse_gold_recall_data).astype(float)
mouse_data_recall = [np.copy(mouse_auto_recall_data), np.copy(mouse_bronze_recall_data), np.copy(mouse_silver_recall_data), np.copy(mouse_gold_recall_data)]
# 计算标准误差
# std_errs = [stats.sem(group) for group in data]
mouse_data_avg_recall = [np.mean(mouse_auto_recall_data), np.mean(mouse_bronze_recall_data), np.mean(mouse_silver_recall_data), np.mean(mouse_gold_recall_data)]
mouse_incr_data_avg_recall = [np.mean(mouse_auto_recall_data), np.mean(mouse_bronze_recall_data) - np.mean(mouse_auto_recall_data), np.mean(mouse_silver_recall_data) - np.mean(mouse_bronze_recall_data), np.mean(mouse_gold_recall_data) - np.mean(mouse_silver_recall_data)]
print(mouse_data_avg_recall)

# log_time_human_normalize_precision = []
# for i in range(len(human_incr_data_avg_recall)):
#     log_time_human_normalize_precision.append(human_incr_data_avg_precision[i] / log_data_human_time_avg[i])
# log_time_human_normalize_recall = []
# for i in range(len(human_incr_data_avg_recall)):
#     log_time_human_normalize_recall.append(human_incr_data_avg_recall[i] / log_data_human_time_avg[i])

# log_time_mouse_normalize_precision = []
# for i in range(len(mouse_incr_data_avg_recall)):
#     log_time_mouse_normalize_precision.append(mouse_incr_data_avg_precision[i] / log_data_mouse_time_avg[i])
# log_time_mouse_normalize_recall = []
# for i in range(len(mouse_incr_data_avg_recall)):
#     log_time_mouse_normalize_recall.append(mouse_incr_data_avg_recall[i] / log_data_mouse_time_avg[i])

# print(log_time_mouse_normalize_precision)
# print(log_time_mouse_normalize_recall)

def get_metric(file_path, sheet_name, auto_column, bronze_column, silver_column, gold_column=None, is_human=True):
    # 打开 Excel 文件
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
    auto_data = []
    count = 0
    for i, cell in enumerate(sheet[auto_column][1:]):
        if sheet[name_column][i + 1].value is not None and sheet[name_column][i + 1].value != '':
            if is_human:
                image_number = sheet[name_column][i + 1].value.split('_')[0]
                if image_number in removed_human_neuron_number:
                    continue
            count += 1
            auto_data.append(float(cell.value))
        else:
            break
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(auto_data)

    bronze_data = []
    for i, cell in enumerate(sheet[bronze_column][1:]):
        if sheet[name_column][i + 1].value is not None and sheet[name_column][i + 1].value != '':
            if is_human:
                image_number = sheet[name_column][i + 1].value.split('_')[0]
                if image_number in removed_human_neuron_number:
                    continue
            bronze_data.append(float(cell.value))
        else:
            break
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(bronze_data)

    silver_data = []
    for i, cell in enumerate(sheet[silver_column][1:]):
        if sheet[name_column][i + 1].value is not None and sheet[name_column][i + 1].value != '':
            if is_human:
                image_number = sheet[name_column][i + 1].value.split('_')[0]
                if image_number in removed_human_neuron_number:
                    continue
            silver_data.append(float(cell.value))
        else:
            break
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(silver_data)

    gold_data = []
    if gold_column is not None:
        for i, cell in enumerate(sheet[gold_column][1:]):
            if sheet[name_column][i + 1].value is not None and sheet[name_column][i + 1].value != '':
                if is_human:
                    image_number = sheet[name_column][i + 1].value.split('_')[0]
                    if image_number in removed_human_neuron_number:
                        continue
                gold_data.append(float(cell.value))
            else:
                break
    # column_data = [float(cell.value) for cell in sheet[column][1:]]
    print(gold_data)

    ground_truth_data = [1.0] * count

    return auto_data, bronze_data, silver_data, gold_data, ground_truth_data

synthetic_auto_data, synthetic_bronze_data, synthetic_silver_data, synthetic_gold_data, synthetic_ground_truth_data = get_metric(synthetic_15_file_path, synthetic_15_sheet_name, synthetic_auto_precision_column, synthetic_bronze_precision_column, synthetic_silver_precision_column, synthetic_gold_precision_column, False)
synthetic_auto_recall_data, synthetic_bronze_recall_data, synthetic_silver_recall_data, synthetic_gold_recall_data, ground_truth_recall_data = get_metric(synthetic_15_file_path, synthetic_15_sheet_name, synthetic_auto_recall_column, synthetic_bronze_recall_column, synthetic_silver_recall_column, synthetic_gold_recall_column, False)

assert len(synthetic_auto_data) == len(synthetic_bronze_data) == len(synthetic_silver_data) == len(synthetic_gold_data)

synthetic_auto_data = np.array(synthetic_auto_data).astype(float)
synthetic_bronze_data = np.array(synthetic_bronze_data).astype(float)
synthetic_silver_data = np.array(synthetic_silver_data).astype(float)
synthetic_gold_data = np.array(synthetic_gold_data).astype(float)
synthetic_data = [np.copy(synthetic_auto_data), np.copy(synthetic_bronze_data), np.copy(synthetic_silver_data), np.copy(synthetic_gold_data)]

synthetic_data_avg_precision = [np.mean(synthetic_auto_data), np.mean(synthetic_bronze_data), np.mean(synthetic_silver_data), np.mean(synthetic_gold_data)]
print(synthetic_data_avg_precision)

assert len(synthetic_auto_recall_data) == len(synthetic_bronze_recall_data) == len(synthetic_silver_recall_data) == len(synthetic_gold_recall_data)

synthetic_auto_recall_data = np.array(synthetic_auto_recall_data).astype(float)
synthetic_bronze_recall_data = np.array(synthetic_bronze_recall_data).astype(float)
synthetic_silver_recall_data = np.array(synthetic_silver_recall_data).astype(float)
synthetic_gold_recall_data = np.array(synthetic_gold_recall_data).astype(float)
synthetic_data_recall = [np.copy(synthetic_auto_recall_data), np.copy(synthetic_bronze_recall_data), np.copy(synthetic_silver_recall_data), np.copy(synthetic_gold_recall_data)]

synthetic_data_avg_recall = [np.mean(synthetic_auto_recall_data), np.mean(synthetic_bronze_recall_data), np.mean(synthetic_silver_recall_data), np.mean(synthetic_gold_recall_data)]
print(synthetic_data_avg_recall)

human_time_effort_precision = []
human_time_effort_recall = []
for i in range(1, len(data_avg_precision)):
    human_time_effort_precision.append(data_avg_precision[i] / data_human_time_avg_accumulate[i] * 60)
for i in range(1, len(data_avg_recall)):
    human_time_effort_recall.append(data_avg_recall[i] / data_human_time_avg_accumulate[i] * 60)

mouse_time_effort_precision = []
mouse_time_effort_recall = []
for i in range(1, len(mouse_data_avg_precision)):
    mouse_time_effort_precision.append(mouse_data_avg_precision[i] / data_mouse_time_avg_accumulate[i] * 60)
for i in range(1, len(mouse_data_avg_recall)):
    mouse_time_effort_recall.append(mouse_data_avg_recall[i] / data_mouse_time_avg_accumulate[i] * 60)

synthetic_time_effort_precision = []
synthetic_time_effort_recall = []
for i in range(1, len(synthetic_data_avg_precision)):
    synthetic_time_effort_precision.append(synthetic_data_avg_precision[i] / data_synthetic_time_avg_accumulate[i] * 60)
for i in range(1, len(synthetic_data_avg_recall)):
    synthetic_time_effort_recall.append(synthetic_data_avg_recall[i] / data_synthetic_time_avg_accumulate[i] * 60)

print("accumulate_time: ", data_synthetic_time_avg_accumulate / 60)

print(human_time_effort_precision)
print(human_time_effort_recall)
print(mouse_time_effort_precision)
print(mouse_time_effort_recall)
print(synthetic_time_effort_precision)
print(synthetic_time_effort_recall)

plt.figure(figsize=(9, 9))

# colors = ['#e99674', '#7db5a2', '#c0c0c0', '#ffe352']
# colors = ['#7db5a2', '#c0c0c0', '#ffe352']
colors = ['#f6f5bc', '#c2bed5', '#8aafc9']
# colors = ['#e4bd95', '#70b092', '#e17477']

# plt.plot(human_time_effort_precision, human_time_effort_recall, linestyle='-', linewidth=3, color='black')
#
# for i in range(len(human_time_effort_precision)):
#     plt.scatter(human_time_effort_precision[i], human_time_effort_recall[i], color=colors[i], edgecolor='black', s=550)

# plt.plot(mouse_time_effort_precision, mouse_time_effort_recall, linestyle='-', linewidth=3, color='black')
#
# for i in range(len(mouse_time_effort_precision)):
#     plt.scatter(mouse_time_effort_precision[i], mouse_time_effort_recall[i], color=colors[i], edgecolor='black', s=660)

plt.plot(synthetic_time_effort_precision, synthetic_time_effort_recall, linestyle='--', linewidth=3, color='black')

for i in range(len(synthetic_time_effort_precision)):
    plt.scatter(synthetic_time_effort_precision[i], synthetic_time_effort_recall[i], color=colors[i], edgecolor='black', s=550)

# for i, stage in enumerate(stages):
#     plt.text(data_avg[i], data_avg_recall[i], stage, fontsize=12, ha='right')

# plt.xlabel('Time effort for precision', fontsize=22)
# plt.ylabel('Time effort for recall', fontsize=22)

plt.xticks(fontsize=32)
plt.yticks(fontsize=32)
# plt.xlim(50, 105)
# plt.ylim(top=105)

plt.tight_layout()
# 移除图形边框
for spine in plt.gca().spines.values():
    if spine.spine_type in ['top', 'right']:
        spine.set_visible(False)

# plt.grid(True)
plt.show()